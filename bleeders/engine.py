"""
bleeders.engine
---------------
Core ETL for the Amazon Bleeders Report.

A "bleeder" = an enabled keyword / target / customer search term that received
MORE THAN `MIN_CLICKS` clicks but generated `MAX_SALES` (zero) sales in the
reporting window.

Public API
----------
    summary, wb = generate(source)      # source = path or file-like .xlsx
    save_workbook(wb, "out.xlsx")       # write to disk
    data = workbook_to_bytes(wb)        # bytes, for a Streamlit download button

Design
------
Columns are matched by HEADER NAME, never by column letter. Amazon reorders
columns between exports; header names are stable. All rules live in the
constants and the BLOCKS table below -- edit config, not code.
"""

from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Rules (edit here if a threshold ever changes)
# ---------------------------------------------------------------------------
MIN_CLICKS = 10          # strictly greater than
MAX_SALES = 0.0          # sales must be <= this (i.e. zero)

# Source column headers in the Amazon bulk export (stable names)
SRC = {
    "state": "State",
    "campaign_state": "Campaign State (Informational only)",
    "clicks": "Clicks",
    "sales": "Sales",
    "campaign_name": "Campaign Name (Informational only)",
    "ad_group_name": "Ad Group Name (Informational only)",
    "customer_search_term": "Customer Search Term",
    "match_type": "Match Type",
    "impressions": "Impressions",
    "ctr": "Click-through Rate",
    "cpc": "CPC",
    "spend": "Spend",
    "acos": "ACOS",
    "roas": "ROAS",
    "orders": "Orders",
    "entity": "Entity",
}
TARGETING_KEYWORD = ["Keyword Text", "Product Targeting Expression"]
TARGETING_DISPLAY = ["Targeting Expression",
                     "Resolved Targeting Expression (Informational only)"]

# One entry per section of the report, in output order.
# `headers` is the exact template header row ("" = the blue operator column).
BLOCKS = [
    {
        "section": "Sponsored Products", "sub": "Search Term",
        "source_tab": "SP Search Term Report",
        "entity_keep": None,
        "targeting_fields": TARGETING_KEYWORD,
        "headers": ["Campaign Name", "Ad Group Name", "Targeting", "Match Type",
                    "", "Customer Search Terms", "Impressions", "Clicks",
                    "Click-Thru Rate (CTR)", "Cost Per Click (CPC)", "Spend",
                    "7 Day Total Sales", "Total Advertising Cost of Sales (ACOS)",
                    "Total Return on Advertising Spend (ROAS)",
                    "7 Day Total Orders (#)"],
    },
    {
        "section": "Sponsored Products", "sub": "Targeting",
        "source_tab": "Sponsored Products Campaigns",
        "entity_keep": {"Keyword", "Product Targeting"},
        "targeting_fields": TARGETING_KEYWORD,
        "headers": ["Campaign Name", "Ad Group Name", "Targeting", "Match Type",
                    "", "Clicks", "Click-Thru Rate (CTR)", "Cost Per Click (CPC)",
                    "Spend", "7 Day Total Sales", "ACOS", "Orders"],
    },
    {
        "section": "Sponsored Brands", "sub": "Search Term",
        "source_tab": "SB Search Term Report",
        "entity_keep": None,
        "targeting_fields": TARGETING_KEYWORD,
        "headers": ["Campaign Name", "Ad Group Name", "Targeting", "Match Type",
                    "", "Customer Search Term", "Clicks", "Click-Thru Rate (CTR)",
                    "Cost Per Click (CPC)", "Spend", "14 Day Total Sales", "ACOS",
                    "Orders"],
    },
    {
        "section": "Sponsored Brands", "sub": "Keywords",
        "source_tab": "Sponsored Brands Campaigns",
        "entity_keep": {"Keyword"},
        "targeting_fields": TARGETING_KEYWORD,
        "headers": ["Campaign Name", "Targeting", "Match Type", "Clicks", "",
                    "Cost Per Click (CPC)", "Spend", "14 Day Total Sales", "ACOS",
                    "Orders"],
    },
    {
        "section": "Sponsored Display", "sub": "Targeting",
        "source_tab": "Sponsored Display Campaigns",
        "entity_drop": {"Ad Group", "Ad group"},   # SOP: "uncheck Ad group"
        "entity_keep": None,
        "targeting_fields": TARGETING_DISPLAY,
        "headers": ["Campaign Name", "Ad Group Name", "Targeting", "Clicks", "",
                    "Spend", "Cost Per Click (CPC)", "14 Day Total Sales", "ACOS",
                    "Orders"],
    },
]

# ---------------------------------------------------------------------------
# Value resolution
# ---------------------------------------------------------------------------
def _first_present(row, fields):
    for f in fields:
        if f in row and pd.notna(row[f]) and str(row[f]).strip() != "":
            return row[f]
    return ""


def _num(v):
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0
        return float(v)
    except (TypeError, ValueError):
        return v if v is not None else ""


def resolve(header, row, block):
    h = header.lower()
    if header == "":
        return ""                                   # blue operator column
    if "customer search term" in h:
        return row.get(SRC["customer_search_term"], "")
    if h == "targeting":
        return _first_present(row, block["targeting_fields"])
    if "match type" in h:
        v = row.get(SRC["match_type"], "")
        return "" if pd.isna(v) else v
    if "ad group name" in h:
        return row.get(SRC["ad_group_name"], "")
    if "campaign name" in h:
        return row.get(SRC["campaign_name"], "")
    if "impressions" in h:
        return _num(row.get(SRC["impressions"]))
    if "clicks" in h:
        return _num(row.get(SRC["clicks"]))
    if "ctr" in h:
        return _num(row.get(SRC["ctr"]))
    if "cpc" in h or "cost per click" in h:
        return _num(row.get(SRC["cpc"]))
    if "spend" in h:
        return _num(row.get(SRC["spend"]))
    if "sales" in h:
        return _num(row.get(SRC["sales"]))
    if "acos" in h:
        return _num(row.get(SRC["acos"]))
    if "roas" in h:
        return _num(row.get(SRC["roas"]))
    if "orders" in h:
        return _num(row.get(SRC["orders"]))
    return ""


def number_format(header):
    h = header.lower()
    if any(k in h for k in ["impressions", "orders"]) or h == "clicks":
        return "#,##0"
    if "ctr" in h or "acos" in h:
        return "0.00%"
    if any(k in h for k in ["cpc", "cost per click", "spend", "sales"]):
        return '$#,##0.00'
    if "roas" in h:
        return "0.00"
    return None


# ---------------------------------------------------------------------------
# Filtering
# ---------------------------------------------------------------------------
def filter_bleeders(df, block):
    if df.empty:
        return df
    d = df.copy()
    d[SRC["clicks"]] = pd.to_numeric(d.get(SRC["clicks"]), errors="coerce").fillna(0)
    d[SRC["sales"]] = pd.to_numeric(d.get(SRC["sales"]), errors="coerce").fillna(0)

    mask = (d[SRC["clicks"]] > MIN_CLICKS) & (d[SRC["sales"]] <= MAX_SALES)
    if SRC["state"] in d:
        mask &= d[SRC["state"]].astype(str).str.strip().str.lower() == "enabled"
    if SRC["campaign_state"] in d:
        mask &= d[SRC["campaign_state"]].astype(str).str.strip().str.lower() == "enabled"

    keep = block.get("entity_keep")
    drop = block.get("entity_drop")
    if (keep or drop) and SRC["entity"] in d:
        ent = d[SRC["entity"]].astype(str).str.strip()
        if keep:
            mask &= ent.isin(keep)
        if drop:
            mask &= ~ent.str.lower().isin({x.lower() for x in drop})
    return d[mask]


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
_FONT = "Arial"
_SECTION_FILL = PatternFill("solid", fgColor="1F3864")
_SECTION_FONT = Font(name=_FONT, bold=True, size=12, color="FFFFFF")
_SUB_FONT = Font(name=_FONT, bold=True, size=11, color="1F3864")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(name=_FONT, bold=True, size=10)
_BLUE_FILL = PatternFill("solid", fgColor="8EB4E3")
_DATA_FONT = Font(name=_FONT, size=10)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def generate(source):
    """Read an Amazon bulk export (path or file-like) and build the report.

    Returns (summary, workbook) where summary is a list of
    (block_label, bleeder_count) tuples.
    """
    xl = pd.ExcelFile(source)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bleeders"

    summary = []
    r = 1
    for block in BLOCKS:
        tab = block["source_tab"]
        df = pd.read_excel(xl, sheet_name=tab) if tab in xl.sheet_names else pd.DataFrame()
        hits = filter_bleeders(df, block)
        summary.append((f'{block["section"]} - {block["sub"]}', len(hits)))

        headers = block["headers"]
        ncol = len(headers)

        ws.cell(row=r, column=1, value=block["section"]).font = _SECTION_FONT
        for cc in range(1, ncol + 1):
            ws.cell(row=r, column=cc).fill = _SECTION_FILL
        r += 1
        ws.cell(row=r, column=1, value=block["sub"]).font = _SUB_FONT
        r += 1

        blue_col = None
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=i, value=h if h else None)
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
            if h == "":
                cell.fill = _BLUE_FILL
                blue_col = i
            else:
                cell.fill = _HEADER_FILL
        r += 1

        for _, srow in hits.iterrows():
            for i, h in enumerate(headers, start=1):
                val = resolve(h, srow, block)
                cell = ws.cell(row=r, column=i, value=val if val != "" else None)
                cell.font = _DATA_FONT
                cell.border = _BORDER
                fmt = number_format(h)
                if fmt:
                    cell.number_format = fmt
                if i == blue_col:
                    cell.fill = _BLUE_FILL
            r += 1

        for _ in range(2):                     # spare operator rows
            if blue_col:
                ws.cell(row=r, column=blue_col).fill = _BLUE_FILL
            r += 1
        r += 1                                 # spacer between blocks

    widths = {1: 34, 2: 22, 3: 34, 4: 12, 5: 14, 6: 30, 7: 13, 8: 10,
              9: 12, 10: 12, 11: 12, 12: 14, 13: 14, 14: 14, 15: 14}
    for col, w in widths.items():
        ws.column_dimensions[chr(64 + col)].width = w
    ws.freeze_panes = "A1"
    return summary, wb


def save_workbook(wb, path):
    wb.save(path)


def workbook_to_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
