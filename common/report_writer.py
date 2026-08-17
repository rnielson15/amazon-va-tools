"""
common.report_writer
--------------------
Shared workbook builder for the bleeders reports. Takes already-filtered rows
and writes them into the stacked-block layout (dark section header, sub-header,
column header row, data, and the blue operator-decision column).

Kept separate so more than one tool can share it without duplicating styling.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_FONT = "Arial"
_SECTION_FILL = PatternFill("solid", fgColor="1F3864")
_SECTION_FONT = Font(name=_FONT, bold=True, size=12, color="FFFFFF")
_SUB_FONT = Font(name=_FONT, bold=True, size=11, color="1F3864")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(name=_FONT, bold=True, size=10)
_BLUE_FILL = PatternFill("solid", fgColor="8EB4E3")
_DATA_FONT = Font(name=_FONT, size=10)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def number_format(header):
    h = header.lower()
    if any(k in h for k in ["impressions", "orders"]) or h == "clicks":
        return "#,##0"
    if "ctr" in h or "acos" in h:
        return "0.00%"
    if any(k in h for k in ["cpc", "cost per click", "spend", "sales"]):
        return "$#,##0.00"
    if "roas" in h:
        return "0.00"
    return None


def build_report(sections):
    """sections: list of dicts with keys
         section (str), sub (str), headers (list[str]; "" = blue column),
         rows (list[dict] keyed by header -> value)
    Returns an openpyxl Workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bleeders"
    r = 1
    max_cols = 1

    for sec in sections:
        headers = sec["headers"]
        ncol = len(headers)
        max_cols = max(max_cols, ncol)

        ws.cell(row=r, column=1, value=sec["section"]).font = _SECTION_FONT
        for cc in range(1, ncol + 1):
            ws.cell(row=r, column=cc).fill = _SECTION_FILL
        r += 1
        ws.cell(row=r, column=1, value=sec["sub"]).font = _SUB_FONT
        r += 1

        blue_col = None
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=i, value=h if h else None)
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
            if h == "":
                cell.fill = _BLUE_FILL
                blue_col = i
            else:
                cell.fill = _HEADER_FILL
        r += 1

        for row in sec["rows"]:
            for i, h in enumerate(headers, start=1):
                val = row.get(h, "")
                cell = ws.cell(row=r, column=i, value=val if val not in ("", None) else None)
                cell.font = _DATA_FONT
                cell.border = _BORDER
                fmt = number_format(h)
                if fmt:
                    cell.number_format = fmt
                if i == blue_col:
                    cell.fill = _BLUE_FILL
            r += 1

        for _ in range(2):
            if blue_col:
                ws.cell(row=r, column=blue_col).fill = _BLUE_FILL
            r += 1
        r += 1

    widths = {1: 34, 2: 24, 3: 34, 4: 12, 5: 14, 6: 16, 7: 12,
              8: 16, 9: 12, 10: 12, 11: 12, 12: 12, 13: 14, 14: 14, 15: 14}
    from openpyxl.utils import get_column_letter
    for col in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 14)
    ws.freeze_panes = "A1"
    return wb


def workbook_to_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
